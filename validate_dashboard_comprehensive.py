#!/usr/bin/env python3
"""
Comprehensive validation script for the Executive Dashboard.
Checks layout, formatting, charts, data integrity, and visual elements.
"""

import os
import glob
from openpyxl import load_workbook
from openpyxl.chart import PieChart, BarChart
from datetime import datetime

def validate_executive_dashboard():
    """Comprehensive validation of the executive dashboard."""
    
    print("🔍 SAP S/4HANA Executive Dashboard Validation")
    print("=" * 50)
    
    # Find the most recent executive summary file
    output_dir = r"c:\Users\jsw73\Downloads\incident_reporting_system\output"
    pattern = os.path.join(output_dir, "executive_summary_*.xlsx")
    excel_files = glob.glob(pattern)
    
    if not excel_files:
        print("❌ No executive summary files found!")
        return False
        
    # Get the most recent file
    latest_file = max(excel_files, key=os.path.getctime)
    file_name = os.path.basename(latest_file)
    print(f"📁 Checking file: {file_name}")
    print(f"📍 Full path: {latest_file}")
    print(f"📊 File size: {os.path.getsize(latest_file):,} bytes")
    
    try:
        # Load the workbook
        wb = load_workbook(latest_file)
        print(f"✅ Workbook loaded successfully")
        print(f"📋 Found {len(wb.worksheets)} worksheet(s)")
        
        validation_results = {
            'sheets_found': 0,
            'metrics_output_valid': False,
            'executive_dashboard_valid': False,
            'header_section': False,
            'kpi_cards': 0,
            'charts_found': 0,
            'performance_table': False,
            'formatting_valid': False,
            'data_integrity': False
        }
        
        # Check each sheet
        for sheet in wb.worksheets:
            validation_results['sheets_found'] += 1
            print(f"\n📊 Sheet: '{sheet.title}'")
            print(f"   Dimensions: {sheet.max_row} rows × {sheet.max_column} columns")
            
            if sheet.title == "Metrics Output":
                validation_results['metrics_output_valid'] = validate_metrics_output(sheet)
                
            elif sheet.title == "Executive Dashboard":
                dashboard_results = validate_executive_dashboard_sheet(sheet)
                validation_results.update(dashboard_results)
        
        # Summary report
        print_validation_summary(validation_results)
        
        return validation_results['executive_dashboard_valid']
        
    except Exception as e:
        print(f"❌ Error validating dashboard: {e}")
        return False

def validate_metrics_output(sheet):
    """Validate the Metrics Output sheet."""
    print("   🔍 Validating Metrics Output sheet...")
    
    # Check if we have data
    if sheet.max_row < 10:
        print("   ⚠️  Warning: Very few rows in Metrics Output")
        return False
    
    # Check for key columns
    if sheet.max_column < 3:
        print("   ❌ Missing expected columns (need at least 3)")
        return False
    
    # Sample some data to verify content
    sample_metrics = []
    for row in range(2, min(6, sheet.max_row + 1)):
        metric = sheet.cell(row, 1).value
        value = sheet.cell(row, 2).value
        if metric and value:
            sample_metrics.append((metric, value))
    
    print(f"   ✅ Found {len(sample_metrics)} sample metrics:")
    for metric, value in sample_metrics[:3]:
        print(f"      • {metric}: {value}")
    
    return len(sample_metrics) > 0

def validate_executive_dashboard_sheet(sheet):
    """Comprehensive validation of the Executive Dashboard sheet."""
    print("   🔍 Validating Executive Dashboard sheet...")
    
    results = {
        'executive_dashboard_valid': False,
        'header_section': False,
        'kpi_cards': 0,
        'charts_found': 0,
        'performance_table': False,
        'formatting_valid': False,
        'data_integrity': False
    }
    
    # 1. Check Header Section
    header_cell = sheet['A1']
    if header_cell.value and 'SAP S/4HANA' in str(header_cell.value):
        results['header_section'] = True
        print("   ✅ Header section found with SAP branding")
        
        # Check if header is merged (indicates proper formatting)
        if sheet.merged_cells:
            print(f"   ✅ Found {len(sheet.merged_cells.ranges)} merged cell ranges")
        else:
            print("   ⚠️  No merged cells found - layout may be basic")
    else:
        print("   ❌ Header section missing or incorrect")
    
    # 2. Check KPI Cards Area (rows 4-9)
    kpi_values = []
    for row in range(4, 10):
        for col in range(1, 17):  # Check across all columns
            cell = sheet.cell(row, col)
            if cell.value and str(cell.value).isdigit():
                kpi_values.append((row, col, cell.value))
    
    results['kpi_cards'] = len(kpi_values)
    if results['kpi_cards'] >= 4:
        print(f"   ✅ Found {results['kpi_cards']} KPI values in card area")
        # Show sample KPI values
        for i, (row, col, value) in enumerate(kpi_values[:4]):
            print(f"      • KPI {i+1}: {value} (at {chr(64+col)}{row})")
    else:
        print(f"   ⚠️  Only found {results['kpi_cards']} KPI values (expected 4+)")
    
    # 3. Check for Charts
    if hasattr(sheet, '_charts') and sheet._charts:
        results['charts_found'] = len(sheet._charts)
        print(f"   ✅ Found {results['charts_found']} chart(s)")
        
        for i, chart in enumerate(sheet._charts):
            chart_type = type(chart).__name__
            print(f"      • Chart {i+1}: {chart_type}")
            if hasattr(chart, 'anchor'):
                print(f"        Position: {chart.anchor}")
    else:
        print("   ⚠️  No charts detected in dashboard")
    
    # 4. Check Performance Table Area (rows 22-30)
    table_headers = []
    for col in range(1, 6):  # Check first 5 columns
        cell = sheet.cell(24, col)  # Row 24 should be table header
        if cell.value:
            table_headers.append(str(cell.value))
    
    if len(table_headers) >= 3:
        results['performance_table'] = True
        print(f"   ✅ Performance table found with headers: {', '.join(table_headers[:3])}")
    else:
        print("   ⚠️  Performance table headers not found")
    
    # 5. Check Data Integrity
    non_empty_cells = 0
    total_cells = 0
    
    for row in range(1, min(31, sheet.max_row + 1)):  # Check first 30 rows
        for col in range(1, min(17, sheet.max_column + 1)):  # Check first 16 columns
            total_cells += 1
            cell = sheet.cell(row, col)
            if cell.value is not None and str(cell.value).strip():
                non_empty_cells += 1
    
    data_density = (non_empty_cells / total_cells) * 100 if total_cells > 0 else 0
    results['data_integrity'] = data_density > 15  # At least 15% of cells should have data
    
    print(f"   📊 Data density: {data_density:.1f}% ({non_empty_cells}/{total_cells} cells)")
    if results['data_integrity']:
        print("   ✅ Data integrity check passed")
    else:
        print("   ⚠️  Low data density - dashboard may be sparse")
    
    # 6. Check Formatting (colors and fonts)
    formatted_cells = 0
    for row in range(1, min(31, sheet.max_row + 1)):
        for col in range(1, min(17, sheet.max_column + 1)):
            cell = sheet.cell(row, col)
            if (cell.font and cell.font.color) or (cell.fill and cell.fill.start_color):
                formatted_cells += 1
    
    results['formatting_valid'] = formatted_cells > 20  # Should have some formatting
    print(f"   🎨 Formatted cells: {formatted_cells}")
    if results['formatting_valid']:
        print("   ✅ Formatting validation passed")
    else:
        print("   ⚠️  Limited formatting detected")
    
    # Overall dashboard validation
    critical_checks = [
        results['header_section'],
        results['kpi_cards'] >= 3,
        results['data_integrity']
    ]
    
    results['executive_dashboard_valid'] = all(critical_checks)
    
    return results

def validate_specific_cells(sheet):
    """Check specific cells that should contain key information."""
    print("   🔍 Validating specific cell content...")
    
    key_cells_to_check = [
        ('A1', 'Main header'),
        ('A3', 'Subtitle/timestamp'),
        ('A4', 'Total Incidents title'),
        ('A5', 'KPI card area'),
        ('A10', 'Chart section'),
        ('A22', 'Performance table section')
    ]
    
    for cell_ref, description in key_cells_to_check:
        cell = sheet[cell_ref]
        if cell.value:
            print(f"      ✅ {cell_ref} ({description}): {str(cell.value)[:50]}")
        else:
            print(f"      ❌ {cell_ref} ({description}): Empty")

def print_validation_summary(results):
    """Print a comprehensive validation summary."""
    print("\n" + "="*50)
    print("📋 VALIDATION SUMMARY")
    print("="*50)
    
    print(f"📊 Sheets found: {results['sheets_found']}/2")
    print(f"✅ Metrics Output valid: {'Yes' if results['metrics_output_valid'] else 'No'}")
    print(f"🎯 Executive Dashboard valid: {'Yes' if results['executive_dashboard_valid'] else 'No'}")
    
    if results['executive_dashboard_valid']:
        print("\n🎉 EXECUTIVE DASHBOARD DETAILS:")
        print(f"   📋 Header section: {'✅ Valid' if results['header_section'] else '❌ Missing'}")
        print(f"   📊 KPI cards found: {results['kpi_cards']}")
        print(f"   📈 Charts found: {results['charts_found']}")
        print(f"   📝 Performance table: {'✅ Found' if results['performance_table'] else '❌ Missing'}")
        print(f"   🎨 Formatting: {'✅ Good' if results['formatting_valid'] else '⚠️ Basic'}")
        print(f"   📋 Data integrity: {'✅ Good' if results['data_integrity'] else '⚠️ Sparse'}")
        
        # Overall score
        total_checks = 6
        passed_checks = sum([
            results['header_section'],
            results['kpi_cards'] >= 4,
            results['charts_found'] > 0,
            results['performance_table'],
            results['formatting_valid'],
            results['data_integrity']
        ])
        
        score = (passed_checks / total_checks) * 100
        print(f"\n🎯 Overall Dashboard Score: {score:.1f}% ({passed_checks}/{total_checks})")
        
        if score >= 80:
            print("🌟 EXCELLENT - Professional dashboard ready for executives!")
        elif score >= 60:
            print("👍 GOOD - Dashboard functional with minor improvements needed")
        else:
            print("⚠️ NEEDS WORK - Dashboard requires attention")
    
    else:
        print("\n❌ Dashboard validation failed - please check the implementation")

def main():
    """Main validation function."""
    print("Starting Executive Dashboard Validation...")
    print(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    success = validate_executive_dashboard()
    
    if success:
        print("\n🎉 Dashboard validation completed successfully!")
        print("The executive dashboard is ready for presentation.")
    else:
        print("\n❌ Dashboard validation failed.")
        print("Please check the implementation and try again.")
    
    return success

if __name__ == "__main__":
    main()