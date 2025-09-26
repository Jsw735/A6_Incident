#!/usr/bin/env python3
"""
Excel Utilities for SAP Incident Analysis
Provides Excel export and formatting capabilities with proper error handling.
"""

import pandas as pd
import numpy as np
from datetime import datetime
from typing import Dict, List, Any, Optional
import logging
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class ExcelUtils:
    """
    Excel utility class following Python best practices.
    Implements proper mathematical operations and error handling.
    """
    
    def __init__(self):
        """Initialize the Excel Utilities."""
        self.logger = logging.getLogger(__name__)
        
        if not OPENPYXL_AVAILABLE:
            self.logger.warning("openpyxl not available - Excel features limited")
    
    def create_executive_report(self, report_data: Dict[str, Any], filename: str) -> bool:
        """
        Create executive summary Excel report using mathematical operations.
        
        Args:
            report_data: Executive report data
            filename: Output filename
            
        Returns:
            True if successful, False otherwise
        """
        try:
            if not OPENPYXL_AVAILABLE:
                return self._create_basic_csv(report_data, filename)
            
            # Create workbook with multiple sheets
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create summary sheet
            self._create_summary_sheet(wb, report_data)
            
            # Create metrics sheet
            self._create_metrics_sheet(wb, report_data)
            
            # Create recommendations sheet
            self._create_recommendations_sheet(wb, report_data)
            
            # Save workbook
            wb.save(filename)
            self.logger.info(f"Executive report saved to {filename}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error creating executive report: {e}")
            return False
    
    def _create_summary_sheet(self, workbook, report_data: Dict[str, Any]):
        """Create executive summary sheet with formatting."""
        ws = workbook.create_sheet("Executive Summary", 0)
        
        # Title
        ws['A1'] = "SAP Incident Management - Executive Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        
        # Merge title cells
        ws.merge_cells('A1:D1')
        
        # Generation timestamp
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].font = Font(italic=True)
        
        # Key Performance Indicators using mathematical operations
        kpis = report_data.get('kpis', {})
        
        row = 4
        ws[f'A{row}'] = "Key Performance Indicators"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 2
        
        # Display KPIs with calculations
        kpi_items = [
            ("Total Incidents", kpis.get('total_incidents', 0)),
            ("Open Incidents", kpis.get('total_open', 0)),
            ("High Priority Open", kpis.get('high_priority_open', 0)),
            ("Average Resolution (hours)", kpis.get('avg_resolution_hours', 0)),
            ("SLA Compliance (%)", kpis.get('sla_compliance', 0)),
            ("Weekly Volume", kpis.get('weekly_volume', 0))
        ]
        
        for label, value in kpi_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            
            # Color coding based on values using comparison operations
            if "SLA Compliance" in label:
                if value >= 95:
                    ws[f'B{row}'].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif value >= 85:
                    ws[f'B{row}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                else:
                    ws[f'B{row}'].fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            row += 1
        
        # Health Score calculation using mathematical operations
        row += 2
        health_score = self._calculate_health_score(kpis)
        ws[f'A{row}'] = "Overall Health Score"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        ws[f'B{row}'] = f"{health_score}/100"
        
        # Color code health score
        if health_score >= 90:
            ws[f'B{row}'].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        elif health_score >= 70:
            ws[f'B{row}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        else:
            ws[f'B{row}'].fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_metrics_sheet(self, workbook, report_data: Dict[str, Any]):
        """Create detailed metrics sheet."""
        ws = workbook.create_sheet("Detailed Metrics")
        
        # Headers
        headers = ["Metric", "Current Value", "Target", "Status", "Trend"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Populate metrics data using mathematical operations
        kpis = report_data.get('kpis', {})
        
        metrics_data = [
            ("Open Incident Rate (%)", kpis.get('open_percentage', 0), 20, "Target"),
            ("High Priority Count", kpis.get('high_priority_open', 0), 5, "Count"),
            ("Resolution Time (hours)", kpis.get('avg_resolution_hours', 0), 24, "Hours"),
            ("SLA Compliance (%)", kpis.get('sla_compliance', 0), 95, "Percentage"),
            ("Weekly Volume", kpis.get('weekly_volume', 0), 50, "Count")
        ]
        
        for row, (metric, current, target, unit) in enumerate(metrics_data, 2):
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=current)
            ws.cell(row=row, column=3, value=target)
            
            # Calculate status using comparison operations
            if unit == "Percentage" and current >= target:
                status = "✓ Good"
                status_color = "90EE90"
            elif unit == "Hours" and current <= target:
                status = "✓ Good"
                status_color = "90EE90"
            elif unit == "Count" and current <= target:
                status = "✓ Good"
                status_color = "90EE90"
            elif unit == "Target" and current <= target:
                status = "✓ Good"
                status_color = "90EE90"
            else:
                status = "⚠ Needs Attention"
                status_color = "FFFF99"
            
            status_cell = ws.cell(row=row, column=4, value=status)
            status_cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
            
            # Simple trend indicator
            ws.cell(row=row, column=5, value="→ Stable")
    
    def _create_recommendations_sheet(self, workbook, report_data: Dict[str, Any]):
        """Create recommendations sheet."""
        ws = workbook.create_sheet("Recommendations")
        
        # Title
        ws['A1'] = "Action Items and Recommendations"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Headers
        ws['A3'] = "Priority"
        ws['B3'] = "Recommendation"
        ws['C3'] = "Expected Impact"
        
        for cell in [ws['A3'], ws['B3'], ws['C3']]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Add recommendations
        recommendations = report_data.get('recommendations', [])
        
        for row, recommendation in enumerate(recommendations[:10], 4):  # Top 10 recommendations
            # Assign priority based on keywords using string operations
            if any(word in recommendation.lower() for word in ['critical', 'urgent', 'immediate']):
                priority = "High"
                priority_color = "FFB6C1"
            elif any(word in recommendation.lower() for word in ['consider', 'plan', 'review']):
                priority = "Medium"
                priority_color = "FFFF99"
            else:
                priority = "Low"
                priority_color = "90EE90"
            
            priority_cell = ws.cell(row=row, column=1, value=priority)
            priority_cell.fill = PatternFill(start_color=priority_color, end_color=priority_color, fill_type="solid")
            
            ws.cell(row=row, column=2, value=recommendation)
            ws.cell(row=row, column=3, value="Improved efficiency")
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 20
    
    def _calculate_health_score(self, kpis: Dict[str, Any]) -> int:
        """Calculate overall health score using mathematical operations."""
        try:
            score = 100  # Start with perfect score
            
            # Deduct points based on performance using subtraction
            open_percentage = kpis.get('open_percentage', 0)
            if open_percentage > 30:
                score -= 25  # Subtraction operation
            elif open_percentage > 20:
                score -= 10
            
            # SLA compliance impact using comparison and subtraction
            sla_compliance = kpis.get('sla_compliance', 100)
            if sla_compliance < 90:
                score -= 20
            elif sla_compliance < 95:
                score -= 10
            
            # High priority incidents impact
            high_priority = kpis.get('high_priority_open', 0)
            if high_priority > 10:
                score -= 15
            elif high_priority > 5:
                score -= 5
            
            return max(0, score)  # Ensure non-negative using comparison
            
        except Exception:
            return 50  # Default middle score if calculation fails
    
    def _create_basic_csv(self, report_data: Dict[str, Any], filename: str) -> bool:
        """Create basic CSV export when openpyxl is not available."""
        try:
            # Convert filename to CSV
            csv_filename = filename.replace('.xlsx', '.csv').replace('.xls', '.csv')
            
            # Create simple summary data
            summary_data = []
            kpis = report_data.get('kpis', {})
            
            for key, value in kpis.items():
                summary_data.append({'Metric': key, 'Value': value})
            
            # Add recommendations
            recommendations = report_data.get('recommendations', [])
            for i, rec in enumerate(recommendations[:5]):
                summary_data.append({'Metric': f'Recommendation_{i+1}', 'Value': rec})
            
            # Create DataFrame and save
            df = pd.DataFrame(summary_data)
            df.to_csv(csv_filename, index=False)
            
            self.logger.info(f"Basic CSV report saved to {csv_filename}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error creating CSV report: {e}")
            return False
    
    def export_data_with_formatting(self, df: pd.DataFrame, filename: str, 
                                  sheet_name: str = "Data") -> bool:
        """
        Export DataFrame to Excel with basic formatting.
        
        Args:
            df: DataFrame to export
            filename: Output filename
            sheet_name: Sheet name
            
        Returns:
            True if successful, False otherwise
        """
        try:
            if not OPENPYXL_AVAILABLE:
                # Fallback to CSV
                csv_filename = filename.replace('.xlsx', '.csv').replace('.xls', '.csv')
                df.to_csv(csv_filename, index=False)
                return True
            
            # Create workbook and add data
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Add DataFrame to worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Format headers
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            # Auto-adjust column widths using mathematical operations
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        # Calculate max length using comparison operations
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set width using addition operation
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(filename)
            return True
            
        except Exception as e:
            self.logger.error(f"Error exporting data: {e}")
            return False