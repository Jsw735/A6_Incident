#!/usr/bin/env python3
"""
Quick script to verify the Executive Dashboard sheets are present in the latest Excel file.
"""

import os
import glob
from openpyxl import load_workbook

def check_latest_executive_summary():
    """Check the latest executive summary file for our expected sheets."""
    
    # Find the most recent executive summary file
    output_dir = r"c:\Users\jsw73\Downloads\incident_reporting_system\output"
    pattern = os.path.join(output_dir, "executive_summary_*.xlsx")
    excel_files = glob.glob(pattern)
    
    if not excel_files:
        print("No executive summary files found!")
        return
        
    # Get the most recent file
    latest_file = max(excel_files, key=os.path.getctime)
    print(f"Checking file: {os.path.basename(latest_file)}")
    
    try:
        # Load the workbook
        wb = load_workbook(latest_file)
        print(f"Found {len(wb.worksheets)} worksheet(s):")
        
        for sheet in wb.worksheets:
            print(f"  - Sheet: '{sheet.title}'")
            print(f"    Dimensions: {sheet.max_row} rows x {sheet.max_column} columns")
            
            # Check if this is our Executive Dashboard
            if sheet.title == "Executive Dashboard":
                print("    ✅ Executive Dashboard sheet found!")
                
                # Check for key elements
                dashboard_title = sheet['B1'].value
                if dashboard_title and "Executive Dashboard" in str(dashboard_title):
                    print(f"    ✅ Dashboard title found: {dashboard_title}")
                    
                # Check for KPI sections
                kpi_sections = ['B3', 'B6', 'B11', 'B22', 'B34']
                for cell_ref in kpi_sections:
                    value = sheet[cell_ref].value
                    if value:
                        print(f"    ✅ Section at {cell_ref}: {value}")
                        
            elif sheet.title == "Metrics Output":
                print("    ✅ Metrics Output sheet found!")
                
    except Exception as e:
        print(f"Error reading file: {e}")

if __name__ == "__main__":
    check_latest_executive_summary()