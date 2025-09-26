#!/usr/bin/env python3
"""
Visual inspection script to display the actual content of the Executive Dashboard.
"""

import os
import glob
from openpyxl import load_workbook

def inspect_dashboard_content():
    """Display the actual content of the Executive Dashboard for visual verification."""
    
    print("🔍 EXECUTIVE DASHBOARD CONTENT INSPECTION")
    print("=" * 60)
    
    # Find the latest file
    output_dir = r"c:\Users\jsw73\Downloads\incident_reporting_system\output"
    pattern = os.path.join(output_dir, "executive_summary_*.xlsx")
    excel_files = glob.glob(pattern)
    
    if not excel_files:
        print("❌ No files found!")
        return
    
    latest_file = max(excel_files, key=os.path.getctime)
    wb = load_workbook(latest_file)
    
    # Focus on Executive Dashboard sheet
    if 'Executive Dashboard' not in [sheet.title for sheet in wb.worksheets]:
        print("❌ Executive Dashboard sheet not found!")
        return
    
    dashboard = wb['Executive Dashboard']
    
    print(f"📊 Inspecting: {os.path.basename(latest_file)}")
    print(f"📐 Dashboard dimensions: {dashboard.max_row} rows × {dashboard.max_column} columns")
    
    # Header Section (Rows 1-4)
    print("\n🏆 HEADER SECTION (Rows 1-4)")
    print("-" * 40)
    for row in range(1, 5):
        for col in range(1, min(5, dashboard.max_column + 1)):
            cell = dashboard.cell(row, col)
            if cell.value:
                cell_ref = f"{chr(64+col)}{row}"
                print(f"  {cell_ref}: {str(cell.value)[:60]}")
    
    # KPI Cards Area (Rows 4-9)
    print("\n📊 KPI CARDS AREA (Rows 4-9)")
    print("-" * 40)
    kpi_data = []
    for row in range(4, 10):
        row_data = []
        for col in range(1, min(17, dashboard.max_column + 1)):
            cell = dashboard.cell(row, col)
            if cell.value is not None:
                row_data.append(f"{chr(64+col)}{row}:{cell.value}")
        if row_data:
            kpi_data.extend(row_data[:4])  # Show first 4 values per row
    
    for i, item in enumerate(kpi_data[:8]):  # Show first 8 KPI items
        print(f"  {item}")
    
    # Charts Area Info
    print("\n📈 CHARTS SECTION")
    print("-" * 40)
    if hasattr(dashboard, '_charts') and dashboard._charts:
        for i, chart in enumerate(dashboard._charts):
            chart_type = type(chart).__name__
            anchor_info = "Unknown position"
            if hasattr(chart, 'anchor') and hasattr(chart.anchor, '_from'):
                col = chart.anchor._from.col
                row = chart.anchor._from.row
                anchor_info = f"{chr(65+col)}{row+1}"
            print(f"  Chart {i+1}: {chart_type} at {anchor_info}")
            
            # Try to get chart title
            if hasattr(chart, 'title') and chart.title:
                print(f"    Title: {chart.title}")
    else:
        print("  No charts found")
    
    # Priority Data Area (Rows 10-16)
    print("\n🎯 PRIORITY DATA SECTION (Rows 12-16)")
    print("-" * 40)
    for row in range(12, 17):
        row_items = []
        for col in range(1, 5):
            cell = dashboard.cell(row, col)
            if cell.value:
                row_items.append(str(cell.value))
        if row_items:
            print(f"  Row {row}: {' | '.join(row_items)}")
    
    # Weekly Trends Data (Rows 12-16, Columns J-M)
    print("\n📅 WEEKLY TRENDS DATA (Columns J-M)")
    print("-" * 40)
    for row in range(12, 17):
        row_items = []
        for col in range(10, 14):  # J, K, L, M columns
            cell = dashboard.cell(row, col)
            if cell.value:
                row_items.append(str(cell.value))
        if row_items:
            print(f"  Row {row}: {' | '.join(row_items)}")
    
    # Performance Table (Rows 24-30)
    print("\n📈 PERFORMANCE TABLE (Rows 24-30)")
    print("-" * 40)
    for row in range(24, 31):
        row_items = []
        for col in range(1, 6):
            cell = dashboard.cell(row, col)
            if cell.value:
                row_items.append(str(cell.value)[:20])  # Truncate long values
        if row_items:
            print(f"  Row {row}: {' | '.join(row_items)}")
    
    # Cell Formatting Sample
    print("\n🎨 FORMATTING SAMPLE")
    print("-" * 40)
    sample_cells = ['A1', 'A5', 'B24', 'C24']
    for cell_ref in sample_cells:
        cell = dashboard[cell_ref]
        formatting_info = []
        
        if cell.font and cell.font.name:
            formatting_info.append(f"Font: {cell.font.name}")
        if cell.font and cell.font.size:
            formatting_info.append(f"Size: {cell.font.size}")
        if cell.font and cell.font.color and hasattr(cell.font.color, 'rgb'):
            formatting_info.append(f"Color: {cell.font.color.rgb}")
        if cell.fill and hasattr(cell.fill, 'start_color') and cell.fill.start_color.rgb != '00000000':
            formatting_info.append(f"Fill: {cell.fill.start_color.rgb}")
        
        if formatting_info:
            print(f"  {cell_ref}: {', '.join(formatting_info)}")
        elif cell.value:
            print(f"  {cell_ref}: No special formatting (Value: {cell.value})")
    
    # Merged Cells Info
    print("\n🔗 MERGED CELLS")
    print("-" * 40)
    if dashboard.merged_cells:
        for i, merged_range in enumerate(list(dashboard.merged_cells.ranges)[:5]):  # Show first 5
            print(f"  Merge {i+1}: {merged_range}")
    else:
        print("  No merged cells found")
    
    print("\n✅ Dashboard content inspection completed!")

if __name__ == "__main__":
    inspect_dashboard_content()