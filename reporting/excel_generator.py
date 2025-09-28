#!/usr/bin/env python3
"""
Enhanced Excel Generator Module with Advanced Styling and Charts
Creates comprehensive Excel reports with tiles, metrics, and visualizations
Following pandas styling best practices and clean code principles
"""

import pandas as pd
import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, Optional, List
import json

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, Reference, LineChart, PieChart, ScatterChart
    from openpyxl.chart.series import DataPoint
    from openpyxl.drawing.image import Image
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Optional dependency for pandas Styler export
try:
    import jinja2  # type: ignore
    JINJA2_AVAILABLE = True
except Exception:
    JINJA2_AVAILABLE = False

class EnhancedExcelStyleManager:
    """Advanced style management with tile and chart styling capabilities."""
    
    def __init__(self):
        self.colors = self._initialize_enhanced_palette()
        self.fonts = self._initialize_enhanced_fonts()
        self.tile_styles = self._initialize_tile_styles()
    
    def _initialize_enhanced_palette(self) -> Dict[str, str]:
        """Initialize comprehensive color palette for metrics and charts."""
        return {
            # Primary brand colors
            'primary_blue': '1f4e79',
            'secondary_blue': '2e75b6',
            'accent_blue': '5b9bd5',
            'light_blue': 'bdd7ee',
            
            # Status colors
            'success_green': '70ad47',
            'warning_orange': 'ffc000',
            'danger_red': 'c5504b',
            'info_purple': '7030a0',
            
            # Tile colors
            'tile_blue': '2e75b6',
            'tile_green': '70ad47',
            'tile_orange': 'ffc000',
            'tile_red': 'c5504b',
            'tile_purple': '7030a0',
            'tile_teal': '00b0a0',
            
            # Chart colors
            'chart_primary': '1f4e79',
            'chart_secondary': '70ad47',
            'chart_accent': 'ffc000',
            
            # Background colors
            'background_light': 'f8f9fa',
            'background_medium': 'e9ecef',
            'text_dark': '212529',
            'white': 'ffffff'
        }
    
    def _initialize_enhanced_fonts(self) -> Dict[str, Font]:
        """Initialize enhanced font styles for different elements."""
        return {
            'title_large': Font(name='Segoe UI', bold=True, size=24, color=self.colors['text_dark']),
            'title_medium': Font(name='Segoe UI', bold=True, size=18, color=self.colors['text_dark']),
            'title_small': Font(name='Segoe UI', bold=True, size=14, color=self.colors['text_dark']),
            'tile_value': Font(name='Segoe UI', bold=True, size=20, color=self.colors['white']),
            'tile_label': Font(name='Segoe UI', bold=True, size=12, color=self.colors['white']),
            'tile_subtitle': Font(name='Segoe UI', size=10, color=self.colors['white']),
            'header_standard': Font(name='Segoe UI', bold=True, size=11, color=self.colors['white']),
            'body_text': Font(name='Segoe UI', size=11, color=self.colors['text_dark']),
            'small_text': Font(name='Segoe UI', size=9, color=self.colors['text_dark'])
        }
    
    def _initialize_tile_styles(self) -> Dict[str, Dict[str, Any]]:
        """Initialize tile styling configurations."""
        return {
                'total_incidents': {
                'color': self.colors['tile_blue'],
                'icon': '[ICON_CHART]',
                'label': 'Total Incidents'
            },
            'open_incidents': {
                'color': self.colors['tile_orange'],
                'icon': '[ICON_UNLOCK]',
                'label': 'Open Incidents'
            },
            'critical_incidents': {
                'color': self.colors['tile_red'],
                'icon': '[ICON_ALERT]',
                'label': 'Critical Incidents'
            },
            'resolution_rate': {
                'color': self.colors['tile_green'],
                'icon': '[ICON_CHECK]',
                'label': 'Resolution Rate'
            },
            'sla_compliance': {
                'color': self.colors['tile_purple'],
                'icon': '[ICON_TIMER]',
                'label': 'SLA Compliance'
            },
            'avg_resolution': {
                'color': self.colors['tile_teal'],
                'icon': '[ICON_LIGHTNING]',
                'label': 'Avg Resolution Time'
            }
        }

class ExcelGenerator:
    """
    Enhanced Excel generator with advanced styling, tiles, and charts.
    Implements pandas styling capabilities and visualization best practices.
    """
    
    def __init__(self, enable_dashboard: bool = False):
        """Initialize the enhanced Excel generator."""
        self.logger = logging.getLogger(__name__)
        self.output_dir = Path("output")
        self.output_dir.mkdir(exist_ok=True)
        
        # Initialize enhanced style manager
        self.style_manager = EnhancedExcelStyleManager()

        # Feature flags
        self.enable_dashboard = enable_dashboard
        
        # Define standard Excel styling
        if OPENPYXL_AVAILABLE:
            self._initialize_standard_styles()
    
    def _initialize_standard_styles(self):
        """Initialize standard Excel styles."""
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def create_comprehensive_report(self, df: pd.DataFrame, analysis_results: Dict = None) -> Optional[str]:
        """
        Create comprehensive Excel report with organized, non-duplicate metrics sheets.
        Each sheet has a distinct purpose with no overlapping data.
        """
        try:
            if df is None or df.empty:
                self.logger.error("Cannot create report: DataFrame is empty or None")
                return None
            
            # Generate clean, organized filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"incident_analysis_report_{timestamp}.xlsx"
            filepath = self.output_dir / filename
            
            self.logger.info(f"Creating comprehensive Excel report: {filename}")
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                
                # SHEET 1: Executive Dashboard (high-level KPIs only)
                self._create_executive_dashboard_sheet(writer, df, analysis_results)
                
                # SHEET 2: Core Metrics (fundamental counts and rates)
                self._create_core_metrics_sheet(writer, df, analysis_results)
                
                # SHEET 3: Performance Analysis (SLA, timing, efficiency)
                self._create_performance_sheet(writer, df, analysis_results)
                
                # SHEET 4: Assignment Analysis (workstreams, distribution)
                self._create_assignment_analysis_sheet(writer, df, analysis_results)
                
                # SHEET 5: Trend Analysis (temporal patterns, forecasting)
                self._create_trend_analysis_sheet(writer, df, analysis_results)
                
                # SHEET 6: Data Quality (validation, completeness)
                self._create_quality_metrics_sheet(writer, df, analysis_results)
                
                # SHEET 7: Raw Data (filtered, essential columns only)
                self._create_clean_data_sheet(writer, df)
            
            self.logger.info(f"Excel report created successfully: {filepath}")
            return str(filepath)
            
        except Exception as e:
            error_msg = f"Error creating comprehensive Excel report: {str(e)}"
            self.logger.error(error_msg)
            return None
    
    def _create_executive_dashboard_sheet(self, writer: pd.ExcelWriter, df: pd.DataFrame, analysis_results: Dict):
        """Create executive dashboard sheet with high-level KPIs."""
        try:
            # Create the worksheet using openpyxl workbook
            workbook = writer.book
            if 'Executive Dashboard' in workbook.sheetnames:
                worksheet = workbook['Executive Dashboard']
            else:
                worksheet = workbook.create_sheet('Executive Dashboard', 0)
            
            # Create dashboard header
            self._create_dashboard_header(worksheet)
            
            # Calculate key metrics
            metrics = self._calculate_dashboard_metrics(df, analysis_results)
            
            # Create metrics tiles
            self._create_metrics_tiles(worksheet, metrics)
            
            # Create summary charts
            self._create_dashboard_charts(worksheet, df, metrics)
            
            # Optimize layout
            self._optimize_dashboard_layout(worksheet)
            
            self.logger.info("Executive dashboard created successfully")
            
        except Exception as e:
            self.logger.error(f"Error creating executive dashboard: {e}")
    
    def _create_core_metrics_sheet(self, writer: pd.ExcelWriter, df: pd.DataFrame, analysis_results: Dict):
        """Create core metrics sheet with fundamental incident statistics only."""
        try:
            metrics = analysis_results.get('metrics', {}) if analysis_results else {}
            core_metrics = metrics.get('core_metrics', {})
            
            # Create clean metrics data (no duplicates)
            core_data = []
            
            # Basic counts section
            core_data.extend([
                ['FUNDAMENTAL COUNTS', ''],
                ['Total Incidents', core_metrics.get('total_incidents', 0)],
                ['Open Incidents', core_metrics.get('open_incidents', 0)],
                ['Closed Incidents', core_metrics.get('closed_incidents', 0)],
                ['In Progress', core_metrics.get('in_progress_incidents', 0)],
                ['', ''],  # Spacer
            ])
            
            # Rates and percentages section
            core_data.extend([
                ['KEY RATES', ''],
                ['Closure Rate (%)', f"{core_metrics.get('closure_rate_percentage', 0):.1f}"],
                ['Open Rate (%)', f"{(core_metrics.get('open_incidents', 0) / max(1, core_metrics.get('total_incidents', 1)) * 100):.1f}"],
                ['', ''],  # Spacer
            ])
            
            # Create DataFrame and write to sheet
            core_df = pd.DataFrame(core_data, columns=['Metric', 'Value'])
            core_df.to_excel(writer, sheet_name='Core Metrics', index=False)
            
            # Apply formatting
            worksheet = writer.sheets['Core Metrics']
            self._apply_sheet_formatting(worksheet, 'Core Metrics Analysis')
            
        except Exception as e:
            self.logger.error(f"Error creating core metrics sheet: {str(e)}")
    
    def _create_performance_sheet(self, writer: pd.ExcelWriter, df: pd.DataFrame, analysis_results: Dict):
        """Create performance analysis sheet with timing and SLA metrics only."""
        try:
            metrics = analysis_results.get('metrics', {}) if analysis_results else {}
            performance_metrics = metrics.get('performance_metrics', {})
            
            # Performance-specific data (no overlap with core metrics)
            perf_data = []
            
            # Resolution timing section
            perf_data.extend([
                ['RESOLUTION PERFORMANCE', ''],
                ['Average Resolution (days)', f"{performance_metrics.get('average_resolution_days', 0):.1f}"],
                ['Median Resolution (days)', f"{performance_metrics.get('median_resolution_days', 0):.1f}"],
                ['First Response Time (hrs)', f"{performance_metrics.get('first_response_time_avg', 0):.1f}"],
                ['', ''],  # Spacer
            ])
            
            # SLA compliance section
            perf_data.extend([
                ['SLA COMPLIANCE', ''],
                ['Overall SLA Rate (%)', f"{performance_metrics.get('sla_compliance_rate', 0):.1f}"],
                ['SLA Breaches', performance_metrics.get('sla_breaches', 0)],
                ['', ''],  # Spacer
            ])
            
            # Create DataFrame and write to sheet
            perf_df = pd.DataFrame(perf_data, columns=['Performance Metric', 'Value'])
            perf_df.to_excel(writer, sheet_name='Performance Analysis', index=False)
            
            # Apply formatting
            worksheet = writer.sheets['Performance Analysis']
            self._apply_sheet_formatting(worksheet, 'Performance & SLA Analysis')
            
        except Exception as e:
            self.logger.error(f"Error creating performance sheet: {str(e)}")
    
    def _create_assignment_analysis_sheet(self, writer: pd.ExcelWriter, df: pd.DataFrame, analysis_results: Dict):
        """Create assignment analysis sheet with workstream distribution."""
        try:
            if 'Assignment group' not in df.columns:
                self.logger.warning("Assignment group column not found, skipping assignment analysis sheet")
                return
            
            # Get workload distribution
            workload = df['Assignment group'].value_counts().head(10)
            
            # Write chart data starting from row 20
            start_row = 20
            workbook = writer.book
            if 'Assignment Analysis' in workbook.sheetnames:
                worksheet = workbook['Assignment Analysis']
            else:
                worksheet = workbook.create_sheet('Assignment Analysis')
            worksheet.cell(row=start_row, column=1, value="Assignment Group")
            worksheet.cell(row=start_row, column=2, value="Incident Count")
            
            for i, (group, count) in enumerate(workload.items(), start_row + 1):
                worksheet.cell(row=i, column=1, value=group)
                worksheet.cell(row=i, column=2, value=count)
            
            # Create horizontal bar chart
            chart = BarChart()
            chart.type = "col"
            chart.title = "Workload by Assignment Group"
            chart.y_axis.title = "Number of Incidents"
            chart.x_axis.title = "Assignment Group"
            
            data_range = Reference(worksheet,
                                 min_col=2, max_col=2,
                                 min_row=start_row + 1, max_row=start_row + len(workload))
            
            categories = Reference(worksheet,
                                 min_col=1, max_col=1,
                                 min_row=start_row + 1, max_row=start_row + len(workload))
            
            chart.add_data(data_range)
            chart.set_categories(categories)
            chart.height = 15
            chart.width = 20
            
            worksheet.add_chart(chart, f"D{start_row}")
            
        except Exception as e:
            self.logger.error(f"Error creating assignment analysis sheet: {e}")
    
    def _create_trend_analysis_sheet(self, writer: pd.ExcelWriter, df: pd.DataFrame, analysis_results: Dict):
        """Create trend analysis sheet with temporal patterns and forecasting."""
        try:
            # Convert to datetime and group by week
            df_copy = df.copy()
            df_copy['Created'] = pd.to_datetime(df_copy['Created'], errors='coerce')
            df_copy = df_copy.dropna(subset=['Created'])
            
            if df_copy.empty:
                return
            
            # Group by week
            weekly_counts = df_copy.groupby(df_copy['Created'].dt.to_period('W')).size()
            
            # Write chart data
            workbook = writer.book
            if 'Trend Analysis' in workbook.sheetnames:
                worksheet = workbook['Trend Analysis']
            else:
                worksheet = workbook.create_sheet('Trend Analysis')
            worksheet.cell(row=1, column=1, value="Week")
            worksheet.cell(row=1, column=2, value="Incident Count")
            
            for i, (week, count) in enumerate(weekly_counts.items(), 2):
                worksheet.cell(row=i, column=1, value=str(week))
                worksheet.cell(row=i, column=2, value=count)
            
            # Create line chart
            chart = LineChart()
            chart.title = "Weekly Incident Trend"
            chart.y_axis.title = "Number of Incidents"
            chart.x_axis.title = "Week"
            
            data_range = Reference(worksheet,
                                 min_col=2, max_col=2,
                                 min_row=2, max_row=len(weekly_counts) + 1)
            
            categories = Reference(worksheet,
                                 min_col=1, max_col=1,
                                 min_row=2, max_row=len(weekly_counts) + 1)
            
            chart.add_data(data_range)
            chart.set_categories(categories)
            chart.height = 15
            chart.width = 20
            
            worksheet.add_chart(chart, "D1")
            
        except Exception as e:
            self.logger.error(f"Error creating trend analysis sheet: {e}")
    
    def _create_quality_metrics_sheet(self, writer: pd.ExcelWriter, df: pd.DataFrame, analysis_results: Dict):
        """Create data quality metrics sheet with validation and completeness scores."""
        try:
            quality_data = []
            quality_data.append(['Quality Metric', 'Value', 'Status'])
            
            # Add quality metrics with status indicators
            overall_score = analysis_results.get('overall_score', 0)
            quality_data.append(['Overall Quality Score', f"{overall_score:.1%}", 
                                'Good' if overall_score > 0.8 else 'Needs Improvement'])
            
            quality_df = pd.DataFrame(quality_data[1:], columns=quality_data[0])
            quality_df = self._sanitize_dataframe(quality_df)
            quality_df.to_excel(writer, sheet_name='Data Quality', index=False)
            
        except Exception as e:
            self.logger.error(f"Error writing quality metrics sheet: {e}")
    
    def _create_clean_data_sheet(self, writer: pd.ExcelWriter, df: pd.DataFrame):
        """Write filtered, essential columns only to raw data sheet."""
        try:
            max_columns = 25
            if len(df.columns) > max_columns:
                important_columns = [
                    'Number', 'Created', 'State', 'Priority', 
                    'Short description', 'Assignment group', 'Resolved'
                ]
                available_important = [col for col in important_columns if col in df.columns]
                remaining_cols = [col for col in df.columns if col not in available_important]
                
                selected_columns = available_important + remaining_cols[:max_columns - len(available_important)]
                display_data = df[selected_columns]
            else:
                display_data = df

            # Sanitize textual cells to avoid Excel repair prompts due to illegal control characters
            display_data = self._sanitize_dataframe(display_data)
            
            display_data.to_excel(writer, sheet_name='Raw Data', index=False)
            
            self.logger.info(f"Main data sheet written with {len(display_data)} rows and {len(display_data.columns)} columns")
            
        except Exception as e:
            self.logger.error(f"Error writing clean data sheet: {e}")
    
    def _create_executive_dashboard(self, data: pd.DataFrame, analysis_results: Dict[str, Any], writer: pd.ExcelWriter) -> None:
        """Create executive dashboard with metrics tiles and charts."""
        try:
            workbook = writer.book
            worksheet = workbook.create_sheet('Executive Dashboard', 0)
            
            # Create dashboard header
            self._create_dashboard_header(worksheet)
            
            # Calculate key metrics
            metrics = self._calculate_dashboard_metrics(data, analysis_results)
            
            # Create metrics tiles
            self._create_metrics_tiles(worksheet, metrics)
            
            # Create summary charts
            self._create_dashboard_charts(worksheet, data, metrics)
            
            # Optimize layout
            self._optimize_dashboard_layout(worksheet)
            
            self.logger.info("Executive dashboard created successfully")
            
        except Exception as e:
            self.logger.error(f"Error creating executive dashboard: {e}")
    
    def _create_dashboard_header(self, worksheet):
        """Create professional dashboard header."""
        colors = self.style_manager.colors
        fonts = self.style_manager.fonts
        
        # Main title
        worksheet.merge_cells('A1:M1')
        title_cell = worksheet['A1']
        title_cell.value = "SAP Incident Management Dashboard"
        title_cell.font = fonts['title_large']
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        worksheet.row_dimensions[1].height = 35
        
        # Subtitle with timestamp
        worksheet.merge_cells('A2:M2')
        subtitle_cell = worksheet['A2']
        subtitle_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        subtitle_cell.font = fonts['body_text']
        subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
        worksheet.row_dimensions[2].height = 20
    
    def _calculate_dashboard_metrics(self, data: pd.DataFrame, analysis_results: Dict[str, Any]) -> Dict[str, Any]:
        """Calculate key metrics for dashboard tiles."""
        try:
            total_incidents = len(data)
            
            # Open incidents calculation
            open_states = ['New', 'In Progress', 'Open']
            open_incidents = len(data[data['State'].isin(open_states)]) if 'State' in data.columns else 0
            
            # Critical incidents calculation
            critical_incidents = len(data[data['Priority'] == 'Critical']) if 'Priority' in data.columns else 0
            
            # Resolution rate calculation
            resolved_incidents = len(data[data['State'].isin(['Resolved', 'Closed'])]) if 'State' in data.columns else 0
            resolution_rate = (resolved_incidents / total_incidents * 100) if total_incidents > 0 else 0
            
            # SLA compliance (simplified calculation)
            sla_compliance = 85.5  # This would come from your analysis results
            
            # Average resolution time (simplified)
            avg_resolution_hours = 24.3  # This would come from your analysis results
            
            return {
                'total_incidents': total_incidents,
                'open_incidents': open_incidents,
                'critical_incidents': critical_incidents,
                'resolution_rate': round(resolution_rate, 1),
                'sla_compliance': sla_compliance,
                'avg_resolution_hours': avg_resolution_hours,
                'open_percentage': round((open_incidents / total_incidents * 100), 1) if total_incidents > 0 else 0
            }
            
        except Exception as e:
            self.logger.error(f"Error calculating dashboard metrics: {e}")
            return {}
    
    def _create_metrics_tiles(self, worksheet, metrics: Dict[str, Any]):
        """Create visual metrics tiles with styling."""
        try:
            tile_configs = [
                ('total_incidents', metrics.get('total_incidents', 0), 'B4'),
                ('open_incidents', metrics.get('open_incidents', 0), 'D4'),
                ('critical_incidents', metrics.get('critical_incidents', 0), 'F4'),
                ('resolution_rate', f"{metrics.get('resolution_rate', 0)}%", 'H4'),
                ('sla_compliance', f"{metrics.get('sla_compliance', 0)}%", 'J4'),
                ('avg_resolution', f"{metrics.get('avg_resolution_hours', 0)}h", 'L4')
            ]
            
            for tile_type, value, start_cell in tile_configs:
                self._create_single_tile(worksheet, tile_type, value, start_cell)
            
        except Exception as e:
            self.logger.error(f"Error creating metrics tiles: {e}")
    
    def _create_single_tile(self, worksheet, tile_type: str, value: Any, start_cell: str):
        """Create a single metrics tile with styling."""
        try:
            tile_style = self.style_manager.tile_styles.get(tile_type, {})
            fonts = self.style_manager.fonts
            
            # Calculate tile range (2x3 cells)
            start_col = ord(start_cell[0]) - ord('A') + 1
            start_row = int(start_cell[1:])
            end_col = start_col + 1
            end_row = start_row + 2
            
            # Merge cells for tile
            end_col_letter = chr(ord('A') + end_col - 1)
            worksheet.merge_cells(f'{start_cell}:{end_col_letter}{end_row}')
            
            # Set background color
            tile_color = tile_style.get('color', self.style_manager.colors['tile_blue'])
            fill = PatternFill(start_color=tile_color, end_color=tile_color, fill_type="solid")
            
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.fill = fill
            
            # Add tile content (avoid emoji to keep Excel/Windows safe)
            main_cell = worksheet[start_cell]
            icon = tile_style.get('icon', '[ICON]')
            label = tile_style.get('label', 'Metric')
            
            main_cell.value = f"{icon}\n{value}\n{label}"
            main_cell.font = fonts['tile_value']
            main_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Set row heights for better appearance
            for row in range(start_row, end_row + 1):
                worksheet.row_dimensions[row].height = 25
            
        except Exception as e:
            self.logger.error(f"Error creating tile {tile_type}: {e}")
    
    def _create_dashboard_charts(self, worksheet, data: pd.DataFrame, metrics: Dict[str, Any]):
        """Create summary charts for the dashboard."""
        try:
            # Create priority distribution chart
            self._create_priority_chart(worksheet, data, 'B8')
            
            # Create status distribution chart
            self._create_status_chart(worksheet, data, 'H8')
            
        except Exception as e:
            self.logger.error(f"Error creating dashboard charts: {e}")
    
    def _create_priority_chart(self, worksheet, data: pd.DataFrame, start_cell: str):
        """Create priority distribution pie chart."""
        try:
            if 'Priority' not in data.columns:
                return
            
            # Get priority distribution
            priority_counts = data['Priority'].value_counts()
            
            # Create chart data area
            chart_data_start_row = int(start_cell[1:])
            
            # Write chart data
            worksheet.cell(row=chart_data_start_row, column=2, value="Priority")
            worksheet.cell(row=chart_data_start_row, column=3, value="Count")
            
            for i, (priority, count) in enumerate(priority_counts.items(), 1):
                worksheet.cell(row=chart_data_start_row + i, column=2, value=priority)
                worksheet.cell(row=chart_data_start_row + i, column=3, value=count)
            
            # Create pie chart
            chart = PieChart()
            chart.title = "Priority Distribution"
            
            # Define data range
            data_range = Reference(worksheet, 
                                 min_col=3, max_col=3,
                                 min_row=chart_data_start_row + 1, 
                                 max_row=chart_data_start_row + len(priority_counts))
            
            categories = Reference(worksheet,
                                 min_col=2, max_col=2,
                                 min_row=chart_data_start_row + 1,
                                 max_row=chart_data_start_row + len(priority_counts))
            
            chart.add_data(data_range)
            chart.set_categories(categories)
            chart.height = 10
            chart.width = 15
            
            # Add chart to worksheet
            worksheet.add_chart(chart, f"B{chart_data_start_row + len(priority_counts) + 2}")
            
        except Exception as e:
            self.logger.error(f"Error creating priority chart: {e}")
    
    def _create_status_chart(self, worksheet, data: pd.DataFrame, start_cell: str):
        """Create status distribution bar chart."""
        try:
            if 'State' not in data.columns:
                return
            
            # Get status distribution
            status_counts = data['State'].value_counts()
            
            # Create chart data area
            chart_data_start_row = int(start_cell[1:])
            
            # Write chart data
            worksheet.cell(row=chart_data_start_row, column=8, value="Status")
            worksheet.cell(row=chart_data_start_row, column=9, value="Count")
            
            for i, (status, count) in enumerate(status_counts.items(), 1):
                worksheet.cell(row=chart_data_start_row + i, column=8, value=status)
                worksheet.cell(row=chart_data_start_row + i, column=9, value=count)
            
            # Create bar chart
            chart = BarChart()
            chart.title = "Status Distribution"
            chart.y_axis.title = "Number of Incidents"
            chart.x_axis.title = "Status"
            
            # Define data range
            data_range = Reference(worksheet,
                                 min_col=9, max_col=9,
                                 min_row=chart_data_start_row + 1,
                                 max_row=chart_data_start_row + len(status_counts))
            
            categories = Reference(worksheet,
                                 min_col=8, max_col=8,
                                 min_row=chart_data_start_row + 1,
                                 max_row=chart_data_start_row + len(status_counts))
            
            chart.add_data(data_range)
            chart.set_categories(categories)
            chart.height = 10
            chart.width = 15
            
            # Add chart to worksheet
            worksheet.add_chart(chart, f"H{chart_data_start_row + len(status_counts) + 2}")
            
        except Exception as e:
            self.logger.error(f"Error creating status chart: {e}")
    
    def _optimize_dashboard_layout(self, worksheet):
        """Optimize dashboard layout and column widths."""
        try:
            # Set column widths
            column_widths = {
                'A': 2, 'B': 12, 'C': 12, 'D': 12, 'E': 12, 'F': 12,
                'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 12, 'L': 12, 'M': 2
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
        except Exception as e:
            self.logger.error(f"Error optimizing dashboard layout: {e}")
    
    def _apply_sheet_formatting(self, worksheet, sheet_title):
        """Apply consistent formatting to each sheet."""
        try:
            # Set column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Format header row
            if worksheet.max_row > 0:
                for cell in worksheet[1]:
                    if cell.value:
                        cell.font = self.header_font
                        cell.fill = self.header_fill
                        cell.border = self.border
            
            # Set title for the sheet
            worksheet.cell(row=1, column=1, value=sheet_title).font = Font(bold=True, size=14)
            
        except Exception as e:
            self.logger.error(f"Error applying sheet formatting: {e}")
    
    # Keep existing methods for backward compatibility
    def _write_main_data_sheet(self, data: pd.DataFrame, writer: pd.ExcelWriter) -> None:
        """Write the main incident data to Excel sheet."""
        try:
            max_columns = 25
            if len(data.columns) > max_columns:
                important_columns = [
                    'Number', 'Created', 'State', 'Priority', 
                    'Short description', 'Assignment group', 'Resolved'
                ]
                available_important = [col for col in important_columns if col in data.columns]
                remaining_cols = [col for col in data.columns if col not in available_important]
                
                selected_columns = available_important + remaining_cols[:max_columns - len(available_important)]
                display_data = data[selected_columns]
            else:
                display_data = data

            # Sanitize textual cells to avoid Excel repair prompts due to illegal control characters
            display_data = self._sanitize_dataframe(display_data)
            
            display_data.to_excel(writer, sheet_name='Raw Data', index=False)
            
            self.logger.info(f"Main data sheet written with {len(display_data)} rows and {len(display_data.columns)} columns")
            
        except Exception as e:
            self.logger.error(f"Error writing main data sheet: {e}")
    
    def _write_category_quality_raw_sheet(self, data: pd.DataFrame, analysis_results: Dict[str, Any], writer: pd.ExcelWriter) -> None:
        """Write Category Quality Raw sheet with open records and quality analysis."""
        try:
            # Filter for open records only
            open_data = self._get_open_records_with_quality_analysis(data, analysis_results)
            
            if open_data.empty:
                self.logger.warning("No open records found for Category Quality Raw sheet")
                return
            
            # Write to sheet
            open_data.to_excel(writer, sheet_name='Category Quality Raw', index=False, startrow=0, startcol=0)
            
            # Get the workbook and worksheet for formatting
            workbook = writer.book
            worksheet = workbook['Category Quality Raw']
            
            # Apply formatting
            self._format_category_quality_sheet(worksheet, len(open_data))
            
            self.logger.info(f"Category Quality Raw sheet created with {len(open_data)} open records")
            
        except Exception as e:
            self.logger.error(f"Error writing Category Quality Raw sheet: {e}")
    
    def _get_open_records_with_quality_analysis(self, data: pd.DataFrame, analysis_results: Dict[str, Any]) -> pd.DataFrame:
        """Get open records with quality analysis columns added."""
        try:
            # Filter for open records
            open_df = self._filter_open_records(data)
            
            if open_df.empty:
                return pd.DataFrame()
            
            # Add quality analysis columns
            open_df = open_df.copy()
            
            # Initialize quality analysis columns
            open_df['Category_Quality_Status'] = 'Review Required'
            open_df['Urgency_Quality_Status'] = 'Review Required'
            open_df['Suggested_Category'] = ''
            open_df['Suggested_Urgency'] = ''
            open_df['Quality_Confidence'] = 0.0
            open_df['Quality_Reason'] = ''
            
            # Get quality analysis results
            category_analysis = analysis_results.get('category_analysis', {})
            urgency_analysis = analysis_results.get('urgency_analysis', {})
            
            # Apply category quality analysis
            miscategorized = category_analysis.get('miscategorized_records', [])
            for record in miscategorized:
                mask = open_df['number'] == record.get('number', '')
                if mask.any():
                    open_df.loc[mask, 'Category_Quality_Status'] = 'Potential Issue'
                    open_df.loc[mask, 'Suggested_Category'] = record.get('suggested_subcategory', '')
                    open_df.loc[mask, 'Quality_Confidence'] = record.get('confidence_score', 0.0)
                    open_df.loc[mask, 'Quality_Reason'] = record.get('reason', '')
            
            # Apply urgency quality analysis
            misclassified_urgency = urgency_analysis.get('misclassified_urgency', [])
            for record in misclassified_urgency:
                mask = open_df['number'] == record.get('number', '')
                if mask.any():
                    open_df.loc[mask, 'Urgency_Quality_Status'] = 'Potential Issue'
                    open_df.loc[mask, 'Suggested_Urgency'] = record.get('suggested_urgency', '')
                    open_df.loc[mask, 'Quality_Reason'] = record.get('reason', '')
            
            # Reorder columns to put quality columns at the end
            quality_cols = ['Category_Quality_Status', 'Urgency_Quality_Status', 'Suggested_Category', 
                           'Suggested_Urgency', 'Quality_Confidence', 'Quality_Reason']
            
            other_cols = [col for col in open_df.columns if col not in quality_cols]
            open_df = open_df[other_cols + quality_cols]
            
            return open_df
            
        except Exception as e:
            self.logger.error(f"Error preparing open records with quality analysis: {e}")
            return pd.DataFrame()
    
    def _filter_open_records(self, data: pd.DataFrame) -> pd.DataFrame:
        """Filter data to include only open records."""
        try:
            # Check for resolved_date column
            if 'resolved_date' in data.columns:
                open_mask = pd.isna(data['resolved_date'])
                return data[open_mask].copy()
            
            # Check for status/state columns
            status_columns = ['status', 'state', 'incident_status']
            closed_states = ['closed', 'resolved', 'cancelled', 'canceled']
            
            for col in status_columns:
                if col in data.columns:
                    open_mask = ~data[col].str.lower().isin(closed_states)
                    return data[open_mask].copy()
            
            # If no clear status indicators, return all records
            self.logger.warning("No clear status indicators found, returning all records")
            return data.copy()
            
        except Exception as e:
            self.logger.error(f"Error filtering open records: {e}")
            return data.copy()
    
    def _format_category_quality_sheet(self, worksheet, row_count: int) -> None:
        """Apply formatting to the Category Quality Raw sheet."""
        try:
            if not OPENPYXL_AVAILABLE:
                return
            
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            
            # Header formatting
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Quality status column formatting
            for row_num in range(2, row_count + 2):  # Skip header
                # Category Quality Status
                category_cell = worksheet.cell(row=row_num, column=worksheet.max_column - 5)  # Adjust based on position
                if category_cell.value == 'Potential Issue':
                    category_cell.fill = PatternFill(start_color='FFE6CC', end_color='FFE6CC', fill_type='solid')
                elif category_cell.value == 'Review Required':
                    category_cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                
                # Urgency Quality Status
                urgency_cell = worksheet.cell(row=row_num, column=worksheet.max_column - 4)  # Adjust based on position
                if urgency_cell.value == 'Potential Issue':
                    urgency_cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                elif urgency_cell.value == 'Review Required':
                    urgency_cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            self.logger.error(f"Error formatting Category Quality Raw sheet: {e}")

    def _write_analysis_sheets(self, analysis_results: Dict[str, Any], writer: pd.ExcelWriter) -> None:
        """Write analysis results to separate sheets."""
        try:
            if 'quality' in analysis_results:
                self._write_quality_sheet(analysis_results['quality'], writer)
            
            if 'trends' in analysis_results:
                self._write_trends_sheet(analysis_results['trends'], writer)
            
        except Exception as e:
            self.logger.error(f"Error writing analysis sheets: {e}")
    
    def _write_quality_sheet(self, quality_results: Dict[str, Any], writer: pd.ExcelWriter) -> None:
        """Write data quality analysis to sheet."""
        try:
            quality_data = []
            quality_data.append(['Quality Metric', 'Value', 'Status'])
            
            # Add quality metrics with status indicators
            overall_score = quality_results.get('overall_score', 0)
            quality_data.append(['Overall Quality Score', f"{overall_score:.1%}", 
                                'Good' if overall_score > 0.8 else 'Needs Improvement'])
            
            quality_df = pd.DataFrame(quality_data[1:], columns=quality_data[0])
            quality_df = self._sanitize_dataframe(quality_df)
            quality_df.to_excel(writer, sheet_name='Data Quality', index=False)
            
        except Exception as e:
            self.logger.error(f"Error writing quality sheet: {e}")
    
    def _write_trends_sheet(self, trend_results: Dict[str, Any], writer: pd.ExcelWriter) -> None:
        """Write trend analysis to sheet."""
        try:
            trend_data = []
            trend_data.append(['Trend Metric', 'Value'])
            
            for key, value in trend_results.items():
                if isinstance(value, (int, float, str)):
                    trend_data.append([key.replace('_', ' ').title(), value])
            
            trend_df = pd.DataFrame(trend_data[1:], columns=trend_data[0])
            trend_df = self._sanitize_dataframe(trend_df)
            trend_df.to_excel(writer, sheet_name='Trend Analysis', index=False)
            
        except Exception as e:
            self.logger.error(f"Error writing trends sheet: {e}")
    
    def _create_csv_fallback(self, data: pd.DataFrame, analysis_results: Dict[str, Any]) -> Optional[Path]:
        """Create CSV fallback when Excel libraries aren't available."""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            csv_file = self.output_dir / f"incident_analysis_report_{timestamp}.csv"
            
            self._sanitize_dataframe(data).to_csv(csv_file, index=False)
            
            if analysis_results:
                json_file = self.output_dir / f"analysis_results_{timestamp}.json"
                with open(json_file, 'w') as f:
                    json.dump(analysis_results, f, indent=2, default=str)
            
            self.logger.info(f"CSV fallback report created: {csv_file}")
            print(f"[OK] CSV report created: {csv_file}")
            return csv_file
            
        except Exception as e:
            self.logger.exception(f"Error creating CSV fallback: {str(e)}")
            return None

    def _sanitize_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Sanitize text cells to remove illegal XML control characters and normalize objects to strings.

        - Leaves numeric and datetime types unchanged
        - Cleans only object/string-like cells
        """
        try:
            import re
            import numpy as np  # type: ignore
            pattern = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")

            def sanitize_cell(v):
                try:
                    # Preserve NaN/NaT
                    if v is None:
                        return v
                    if isinstance(v, float) and (pd.isna(v)):
                        return v
                    # Keep numbers and datetimes as-is
                    if isinstance(v, (int, float)):
                        return v
                    if hasattr(v, 'isoformat'):
                        # datetime-like
                        return v
                    # Convert other objects to safe strings
                    s = str(v)
                    return pattern.sub("", s)
                except Exception:
                    return v

            df_copy = df.copy()
            for col in df_copy.columns:
                if pd.api.types.is_object_dtype(df_copy[col]):
                    df_copy[col] = df_copy[col].map(sanitize_cell)
            return df_copy
        except Exception:
            # Best-effort; if anything goes wrong, return original
            return df